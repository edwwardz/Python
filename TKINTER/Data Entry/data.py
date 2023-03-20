from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

root = Tk()
root.title("Data Entry")
root.geometry("700x400+300+200")
root.resizable(False, False)
root.configure(bg="#326273")

file = pathlib.Path("C:\Users\edwar\Music\My Code\python\TKINTER\Data Entry\data.py""")
if file.exists():
    pass
else:
    file=Workbook()
    sheet = file.active
    sheet['A1'] = "Full Name"
    sheet["B1"] = 'Phone Number'
    sheet["C1"] = "Age"
    sheet["D1"] = "Gender"
    sheet["E1"] = "Address"
    file.save('Backend.xlsx')

def submit():
    name = nameValue.get()
    contact = contactValue.get()
    age = ageValue.get()
    gender = Gender_combobox.get()
    address = addressEntry.get(1.0, END)

    print(name)
    print(contact)
    print(age)
    print(gender)
    print(address)

    file=openpyxl.load_workbook("Backend.xlsx")
    sheet=file.active
    sheet.cell(column=1, row=sheet.max_row+1, value=name)
    sheet.cell(column=2, row=sheet.max_row, value=contact)
    sheet.cell(column=3, row=sheet.max_row, value=age)
    sheet.cell(column=4, row=sheet.max_row, value=gender)
    sheet.cell(column=5, row=sheet.max_row, value=address)

    file.save("Backend.xlsx")

    messagebox.showinfo("info", 'detail add!')

    nameValue.set('')
    contactValue.set('')
    ageValue.set('')
    addressEntry.delete(1.0, END)
    Gender_combobox.set("Male")

def clear():
    nameValue.set('')
    contactValue.set('')
    ageValue.set('')
    addressEntry.delete(1.0, END)
    Gender_combobox.set("Male")


#heading
Label(root, text="Please Fill out this Entry Form", font='arial 13', bg="#326273", fg="#fff").place(x=20, y=20)

#Label
Label(root,text="Name", font=23, bg="#326273", fg='#fff').place(x=50, y=100)
Label(root,text="Handphone", font=23, bg="#326273", fg='#fff').place(x=50, y=150)
Label(root,text="Age", font=23, bg="#326273", fg='#fff').place(x=50, y=200)
Label(root,text="Gender", font=23, bg="#326273", fg='#fff').place(x=380, y=200)
Label(root,text="Address", font=23, bg="#326273", fg='#fff').place(x=50, y=250)

#Entry 
nameValue = StringVar()
contactValue = StringVar()
ageValue = StringVar()

nameEntry = Entry(root, textvariable=nameValue, width=35, bd=2, font=20)
ContactEntry = Entry(root, textvariable=contactValue, width=35, bd=2, font=20)
ageEntry = Entry(root, textvariable=ageValue, width=15, bd=2, font=20)

#Gender
Gender_combobox = Combobox(root, values=['Male', "Female"], font='arial 14', state='r', width=10)
Gender_combobox.place(x=460,y=200)
Gender_combobox.set("Male")

addressEntry = Text(root, width=50, height=4, bd=4)

addressEntry.place(x=200, y = 250)
nameEntry.place(x=200, y =100)
ContactEntry.place(x=200, y=150)
ageEntry.place(x=200, y=200)

Button(root, text="Submit", bg="#326273", fg="White",width=15, height=2, command=submit).place(x=200,y=350)
Button(root, text="Clear", bg="#326273",fg="White", width=15, height=2, command=clear).place(x=340, y=350)
Button(root,text="Exit", bg="#326273",fg="White", width=15, height=2, command=lambda:root.destroy()).place(x=480, y=350)



root.mainloop()